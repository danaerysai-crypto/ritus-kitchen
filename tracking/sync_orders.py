import pandas as pd
import requests
from openpyxl import load_workbook
import sys

# ============================================================================
# CONFIGURATION — Replace with your actual values
# ============================================================================

# Get this from Google Sheets: File → Share → Publish to web → CSV
GOOGLE_SHEET_CSV_URL = "https://docs.google.com/spreadsheets/d/e/YOUR_SHEET_ID/pub?gid=0&single=true&output=csv"

# Or use a local CSV path for testing
LOCAL_CSV_PATH = None  # Set to e.g. "orders_export.csv" if testing locally

# ============================================================================
# COLUMN MAPPING — Match your Google Form questions
# ============================================================================

COLUMN_MAP = {
    "Timestamp": "Date",
    "Your Name": "Customer Name",
    "WhatsApp Phone Number": "Phone",
    "How did you hear about us?": "Source",
    "Deliver to (area / guesthouse)": "Delivery Area",
    "Meal need": "Meal",
    "How many people?": "Quantity",
    "Dietary needs (optional)": "Dietary Note",
    "Who referred you?": "Referral By",
    "Anything else?": "Notes"
}

PRODUCT_PRICES = {
    "Lunch": ("Today's Thali", 220),
    "Dinner": ("Today's Thali", 220),
    "Both": ("Today's Thali", 440),
    "Weekly Tiffin": ("Weekly Tiffin", 1100)
}

OUTPUT_COLUMNS = [
    "Date", "Customer Name", "Phone", "Source", "UTM Campaign",
    "Product", "Quantity", "Price", "Payment Mode", "Delivery Area",
    "Dietary Note", "New or Repeat", "Referral By", "Notes"
]

# ============================================================================
# MAIN SYNC
# ============================================================================

def load_orders():
    if LOCAL_CSV_PATH:
        df = pd.read_csv(LOCAL_CSV_PATH)
    else:
        try:
            response = requests.get(GOOGLE_SHEET_CSV_URL)
            response.raise_for_status()
            df = pd.read_csv(pd.io.common.StringIO(response.text))
        except Exception as e:
            print(f"ERROR: Could not fetch Google Sheet. {e}")
            print("Please check GOOGLE_SHEET_CSV_URL or set LOCAL_CSV_PATH for testing.")
            sys.exit(1)
    return df

def clean_orders(df):
    # Rename columns
    df = df.rename(columns=COLUMN_MAP)
    
    # Convert date
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
    
    # Map meal to product + price
    def map_product(meal):
        return PRODUCT_PRICES.get(meal, (meal, 220))
    
    mapped = df["Meal"].apply(map_product)
    df["Product"] = mapped.apply(lambda x: x[0])
    df["Price"] = mapped.apply(lambda x: x[1])
    
    # Clean quantity
    df["Quantity"] = df["Quantity"].astype(str).str.extract(r"(\d+)").fillna(1).astype(int)
    
    # Fill defaults
    df["UTM Campaign"] = "website_order_form"
    df["Payment Mode"] = "Pending"
    df["New or Repeat"] = "New"
    df["Dietary Note"] = df["Dietary Note"].fillna("None")
    df["Referral By"] = df["Referral By"].fillna("")
    df["Notes"] = df["Notes"].fillna("")
    df["Delivery Area"] = df["Delivery Area"].fillna("")
    
    # Select output columns
    return df[OUTPUT_COLUMNS]

def save_to_excel(df):
    try:
        wb = load_workbook("ritus-kitchen-marketing-dashboard.xlsx")
    except FileNotFoundError:
        print("ERROR: ritus-kitchen-marketing-dashboard.xlsx not found.")
        print("Create it first or run from the tracking/ folder.")
        sys.exit(1)
    
    ws = wb["Orders"]
    
    # Clear old data (keep header row 1)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    
    # Append new rows
    for row in df.itertuples(index=False, name=None):
        ws.append(row)
    
    wb.save("ritus-kitchen-marketing-dashboard.xlsx")
    print(f"✅ Synced {len(df)} orders to Excel dashboard.")

if __name__ == "__main__":
    print("Ritu's Kitchen — Order Sync")
    print("Loading orders from source...")
    df = load_orders()
    print(f"Found {len(df)} raw responses.")
    
    print("Cleaning and mapping data...")
    clean_df = clean_orders(df)
    
    print("Saving to Excel...")
    save_to_excel(clean_df)
    
    print("Done. Open ritus-kitchen-marketing-dashboard.xlsx to review.")
